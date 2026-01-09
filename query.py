import threading
import datetime
import pandas as pd
import io
import sqlite3
from pathlib import Path
from nicegui import ui
from starlette.formparsers import MultiPartParser
from nicegui.elements.upload import UploadEventArguments # type: ignore
from nicegui.elements.upload import MultiUploadEventArguments # type: ignore
import sys
import os
import tempfile

MultiPartParser.spool_max_size = 1024 * 1024 * 50

class Logger:
    def __init__(self):
        self.logs = []
        self.lock = threading.Lock()
        self.callbacks = []

    def __enter__(self):
        if len(self.logs) > 2000:
            self.logs = self.logs[-5:]

    def __exit__(self, exc_type, exc_value, traceback):
        pass

    def _notify(self):
        for cb in self.callbacks:
            try:
                cb(self.return_logs())
            except Exception:
                pass

    def append(self, text: str) -> None:
        timestamp = datetime.datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
        with self.lock:
            self.logs.append(f"{timestamp} {text}")
        self._notify()

    def return_logs(self) -> str:
        return "\n".join(self.logs) if self.logs else ""

    def register_callback(self, cb):
        if cb not in self.callbacks:
            self.callbacks.append(cb)

    def get_all(self) -> None:
        with self.lock:
            with open('log.txt', 'w', encoding='utf-8') as f:
                for line in self.logs:
                    f.write(line + '\n')

    def clear(self) -> None:
        with self.lock:
            self.logs.clear()
        self._notify()

Global_logger = Logger()

class ISBN_Database:
    PATH: str = 'ISBN.db'

    def __init__(self, path: str = '') -> None:
        self.path = path if path else ISBN_Database.PATH
        self.conn = None
        self.support_args = ['single', 'multi']

    def __enter__(self) -> sqlite3.Connection:
        self.conn = sqlite3.connect(self.path)
        self.conn.execute("PRAGMA cache_size = 100000")
        self.conn.execute("PRAGMA temp_store = MEMORY")
        return self.conn

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self.conn:
            try:
                self.conn.close()
            except Exception as e:
                Global_logger.append(f"关闭数据库出错: {e}")

    async def create_database(self, **kwargs) -> None:
        import os
        db_name = None
        if 'single' in kwargs and kwargs['single']:
            file = kwargs['single']
            base = os.path.splitext(file.name)[0]
            db_name = f"{base}.db"
        elif 'multi' in kwargs and kwargs['multi']:
            files = kwargs['multi']
            base = os.path.splitext(files[0].name)[0]
            db_name = f"{base}.db"
        else:
            db_name = ISBN_Database.PATH
        self.path = db_name

        data = await Tools.uni_entry(self, **kwargs)
        data = data.drop_duplicates()
        data.columns = ['ISBN']
        data = data[data['ISBN'] != '标准号']
        Global_logger.append(f"正在创建数据库 {self.path}，请稍候...")
        with sqlite3.connect(self.path) as conn:
            data.to_sql('ISBN_table', conn,
                        if_exists='replace', index=False)
            conn.execute("PRAGMA cache_size = 100000")
            conn.execute("PRAGMA temp_store = MEMORY")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_isbn ON ISBN_table(ISBN)")
            conn.commit()
        Global_logger.append("数据库创建完成。")

    async def update_database(self, **kwargs):
        data = await Tools.uni_entry(self, **kwargs)
        data = data.drop_duplicates()
        isbn_col_candidates = ['ISBN', '标准号', 'ISBN标准号']
        isbn_col = None
        for col in isbn_col_candidates:
            if col in data.columns:
                isbn_col = col
                break
        if not isbn_col:
            Global_logger.append("文件缺少 ISBN/标准号/ISBN标准号 列。")
            return
        df = data[[isbn_col]].copy()
        df = df[df[isbn_col].notnull()]
        df = df[df[isbn_col] != '标准号']
        df[isbn_col] = df[isbn_col].astype(str).str.replace('-', '').str.replace(' ', '')
        df.columns = ['ISBN']
        Global_logger.append("正在更新数据库，请稍候...")
        with sqlite3.connect(self.path) as conn:
            df.to_sql('Temp_ISBN', conn, if_exists='replace', index=False)
            conn.execute("PRAGMA cache_size = 100000")
            conn.execute("PRAGMA temp_store = MEMORY")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_temp_isbn ON Temp_ISBN(ISBN)")
            conn.execute(
                "INSERT OR IGNORE INTO ISBN_table(ISBN) SELECT ISBN FROM Temp_ISBN")
            conn.commit()
            conn.execute("DROP TABLE Temp_ISBN")
        Global_logger.append("数据库更新完成。")

    def is_exist(self) -> bool:
        return Path(self.path).exists()

    def get_all_recs(self) -> int:
        with self.conn as conn:
            cursor = conn.execute("SELECT COUNT(*) FROM ISBN")
            return cursor.fetchone()[0]

class Tools:
    @staticmethod
    async def uni_entry(db: ISBN_Database, **kwargs) -> pd.DataFrame:
        if not any(arg in kwargs for arg in db.support_args):
            Global_logger.append("错误：不支持的参数")
            raise ValueError("不支持的参数")
        data = pd.DataFrame()
        if 'single' in kwargs:
            file = kwargs['single']
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                data = await Tools.handle_xlsx_file(file)
            elif file.name.endswith('.csv'):
                data = await Tools.handle_csv_file(file)
            else:
                Global_logger.append("错误：上传的文件格式不受支持。请上传 .xlsx/.xls 或 .csv 格式的文件。")
                raise ValueError("上传的文件格式不受支持。")
        elif 'multi' in kwargs:
            files = kwargs['multi']
            if all(file.name.endswith('.xlsx') or file.name.endswith('.xls') for file in files):
                data = await Tools.merge_xlsx_files(files)
            elif all(file.name.endswith('.csv') for file in files):
                data = await Tools.merge_csv_files(files)
            else:
                Global_logger.append("错误：上传的文件格式不一致或不受支持。请上传 全部 为 .xlsx/.xls 或 .csv 格式的文件。")
                raise ValueError("上传的文件格式不一致或不受支持。")
        return data

    @staticmethod
    async def merge_xlsx_files(file_list: list) -> pd.DataFrame:
        from openpyxl import load_workbook
        import tempfile
        dfs = []
        for file in file_list:
            content = await file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(content)
                tmp.flush()
                wb = load_workbook(tmp.name, data_only=True)
                ws = wb.active
                data = ws.values
                columns = next(data)
                df = pd.DataFrame(data, columns=columns)
                dfs.append(df)
        merged_df = pd.concat(dfs, ignore_index=True)
        return merged_df

    @staticmethod
    async def handle_xlsx_file(file) -> pd.DataFrame:
        from openpyxl import load_workbook
        import tempfile
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(content)
            tmp.flush()
            wb = load_workbook(tmp.name, data_only=True)
            ws = wb.active
            data = ws.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
        return df

    @staticmethod
    async def handle_csv_file(file) -> pd.DataFrame:
        df = pd.read_csv(io.BytesIO(await file.read()))
        return df

    @staticmethod
    async def merge_csv_files(file_list: list) -> pd.DataFrame:
        dfs = []
        for file in file_list:
            df = pd.read_csv(io.BytesIO(await file.read()))
            dfs.append(df)
        merged_df = pd.concat(dfs, ignore_index=True)
        return merged_df

class SearchEngine:
    @staticmethod
    def search_single(db: sqlite3, ISBN: str):
        isbn = ISBN.replace("-", "").replace(" ", "")
        if not isbn:
            Global_logger.append("非法输入")
            return
        Global_logger.append(f"搜索: {isbn}")
        with db as conn:
            sql = "SELECT 1 FROM ISBN_table WHERE ISBN = ? LIMIT 1"
            cursor = conn.execute(sql, (isbn,))
            exists = cursor.fetchone() is not None
        Global_logger.append(f"{isbn} {'存在' if exists else '不存在'}")

    @staticmethod
    def search_batch(db: sqlite3, ISBN_list: list[str]):
        isbn_in = [s.replace("-", "").replace(" ", "") for s in ISBN_list if s.strip()]
        if not isbn_in:
            Global_logger.append("没有可用的输入。")
            return
        found: set = set()
        batch_size = 500
        with db as conn:
            for i in range(0, len(isbn_in), batch_size):
                batch = isbn_in[i: i + batch_size]
                params = ','.join(['?'] * len(batch))
                sql = f"SELECT ISBN FROM ISBN_table WHERE ISBN IN ({params})"
                cursor = conn.execute(sql, batch)
                found.update(row[0] for row in cursor.fetchall())
        for isbn in isbn_in:
            Global_logger.append(f"{isbn} {'存在' if isbn in found else '不存在'}")

    @staticmethod
    async def batch_clean(db: sqlite3, excel_files):
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import tempfile
        import shutil
        result_isbns = []
        isbn_col_candidates = ['ISBN', '标准号', 'ISBN标准号']
        for file in excel_files:
            Global_logger.append(f"读取 {file.name} 中...")
            if file.name.endswith('.xlsx'):
                content = await file.read()
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(content)
                    tmp.flush()
                    wb = load_workbook(tmp.name)
                    ws = wb.active
                    data = ws.values
                    columns = next(data)
                    df = pd.DataFrame(data, columns=columns)
            elif file.name.endswith('.csv'):
                content = await file.read()
                df = pd.read_csv(io.BytesIO(content), dtype=str)
                wb = None
            else:
                Global_logger.append(f"文件 {file.name} 格式不支持")
                continue
            isbn_col = None
            for col in isbn_col_candidates:
                if col in df.columns:
                    isbn_col = col
                    break
            if not isbn_col:
                Global_logger.append(f"文件 {file.name} 缺少 ISBN/标准号/ISBN标准号 列。")
                continue
            df = df[df[isbn_col].notnull()]
            df = df[df[isbn_col] != '标准号']
            df[isbn_col] = df[isbn_col].astype(str).str.replace('-', '').str.replace(' ', '')
            isbn_list = df[isbn_col].tolist()
            found = set()
            batch_size = 500
            with db as conn:
                for i in range(0, len(isbn_list), batch_size):
                    batch = isbn_list[i:i+batch_size]
                    params = ','.join(['?'] * len(batch))
                    sql = f"SELECT ISBN FROM ISBN_table WHERE ISBN IN ({params})"
                    cursor = conn.execute(sql, batch)
                    found.update(row[0] for row in cursor.fetchall())
            df_cleaned = df[~df[isbn_col].isin(found)]
            file_path = Path(file.name)
            if file.name.endswith('.xlsx') and wb is not None:
                # 清空原有数据，只保留表头
                ws_clean = wb.active
                # 删除除第一行外的所有行
                ws_clean.delete_rows(2, ws_clean.max_row)
                # 将df_cleaned写入
                for r in dataframe_to_rows(df_cleaned, index=False, header=False):
                    ws_clean.append(r)
                # 保留原有样式，保存为新文件
                wb.save(file_path)
            elif file.name.endswith('.csv'):
                df_cleaned.to_csv(file_path, index=False)
            else:
                # 其它格式，默认用openpyxl新建并写入
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                wb_new = Workbook()
                ws_new = wb_new.active
                ws_new.append(df_cleaned.columns.tolist())
                for r in dataframe_to_rows(df_cleaned, index=False, header=False):
                    ws_new.append(r)
                wb_new.save(file_path)
            result_isbns.extend(df_cleaned[isbn_col].dropna().tolist())
        if not result_isbns:
            Global_logger.append("没有可用的上传文件数据。")
        Global_logger.append(f"完成，保留 {len(result_isbns)} 条")

class Aplication:
    def get_db_files(self):
        import os
        db_files = [f for f in os.listdir('.') if f.endswith('.db')]
        return db_files if db_files else ['ISBN.db']

    def __init__(self):
        Global_logger.append("应用启动")
        self.database = ISBN_Database()

    def __enter__(self):
        if not self.database.is_exist():
            self._build_init_ui()
            Global_logger.append("初始化中...请上传文件创建数据库")
        else:
            self._build_main_ui()
            Global_logger.append("数据库已存在，进入主界面")
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        Global_logger.append("应用关闭")

    def _build_init_ui(self) -> None:
        ui.markdown('# **初始化**').style("align-self:center;")
        ui.markdown('## **数据库初始化界面**')
        ui.markdown('### **上传 .xlsx 或 .csv 文件以创建数据库：**')
        ui.upload(
            multiple=True,
            on_upload=self._setup_database,
            on_multi_upload=self._setup_database_multi,
            label='单或多传文件，必须一致格式'
        ).style("font-size:20px;")
        ui.separator().style("width:800px; margin-top:10px; margin-bottom:10px;")
        ui.markdown('## **运行日志**').style("align-self:center;")
        with ui.scroll_area() as scroll_area:
            self.text = ui.markdown('').style("white-space: pre-wrap; font-size:16px;")
            self.text.set_content(Global_logger.return_logs())
            def update_logs(logs):
                self.text.set_content(logs)
                scroll_area.scroll_to(percent=1.0)
            Global_logger.register_callback(update_logs)
            self.text.props('readonly')
        ui.separator().style("flex-grow:1;")
        with ui.row().style('margin-top:20px; justify-content: flex-start; gap: 20px;'):
            ui.button('清空日志', on_click=Global_logger.clear).style('font-size: 16px; padding: 10px 20px;')
            ui.button('导出日志', on_click=Global_logger.get_all).style('font-size: 16px; padding: 10px 20px;')
        ui.separator().style("flex-grow:1;")
        ui.button('关闭应用', on_click=self._close_app).style('position: fixed; bottom: 20px; right: 20px; font-size: 16px; padding: 10px 20px;')
        ui.run(reload=False, title='ISBN查重系统 - 初始化界面')

    def _build_main_ui(self) -> None:
        ui.markdown('# **ISBN 查重系统**').style("align-self:center;")
        db_files = self.get_db_files()
        db_select = ui.select(db_files, value=self.database.path, label='选择数据库文件').style('width:300px; margin-bottom:20px;')
        ui.button('切换数据库', on_click=lambda: self._switch_database(db_select.value)).style('margin-left:10px;')
        ui.separator().style("width:800px; margin-top:10px; margin-bottom:10px;")
        with ui.tabs() as main_tabs:
            tab1 = ui.tab('单次查询').style("align-self:center; font-size: 20px; min-width: 150px;")
            tab2 = ui.tab('批量查询').style("align-self:center; font-size: 20px; min-width: 150px;")
            tab3 = ui.tab('更新数据库').style("align-self:center; font-size: 20px; min-width: 150px;")
            tab4 = ui.tab('新建数据库').style("align-self:center; font-size: 20px; min-width: 150px;")
        with ui.tab_panels(main_tabs, value=tab1):
            with ui.tab_panel(tab1):
                ui.markdown('## **单次**')
                text_in_single = ui.input('请输入 ISBN，可带连字符或空格：', placeholder='978-7-111-12345-6').style("width:400px;")
                ui.button('查询', on_click=lambda: self._on_single_check(text_in_single.value)).style("font-size:22px; margin-left:10px;")
            with ui.tab_panel(tab2):
                ui.markdown('## **批量**')
                with ui.tabs() as sub_tabs:
                    sub_tab1 = ui.tab('输入框输入')
                    sub_tab2 = ui.tab('上传文件')
                with ui.tab_panels(sub_tabs, value=sub_tab1):
                    with ui.tab_panel(sub_tab1):
                        ui.markdown('### **请输入多个 ISBN，每行一个，可带连字符或空格：**')
                        batch_textarea = ui.textarea(placeholder='978-7-111-12345-6\n978 7 111 12345 7\n9787111123458').style("width:600px; height:300px;")
                        ui.button('查询', on_click=lambda _: self._on_batch_check(batch_textarea.value)).style("font-size:22px; margin-top:10px;")
                    with ui.tab_panel(sub_tab2):
                        ui.markdown('### **上传包含 ISBN 列的 .xlsx 或 .csv 文件，去重并返回**')
                        ui.upload(auto_upload=True, on_upload=self._send_to_query).style("font-size:20px;").classes('max-w-full').props('accept=".xlsx,.csv,.xls"')
            with ui.tab_panel(tab3):
                ui.markdown('## **更新数据库**')
                ui.markdown('### **上传 .xlsx 或 .csv 文件，系统自动更新：**')
                ui.upload(auto_upload=True, multiple=True, on_upload=self._send_to_update).style("font-size:20px;")
            with ui.tab_panel(tab4):
                ui.markdown('# **新建数据库**').style("align-self:center;")
                ui.markdown('### **上传 .xlsx 或 .csv 文件以创建数据库：**')
                ui.upload(multiple=True, on_upload=self._setup_database, on_multi_upload=self._setup_database_multi, label='单或多传文件，必须一致格式').style("font-size:20px;")
        ui.separator().style("width:800px; margin-top:10px; margin-bottom:10px;")
        ui.markdown('## **运行日志**').style("align-self:center;")
        with ui.scroll_area() as scroll_area:
            self.text = ui.markdown('').style("white-space: pre-wrap; font-size:16px;")
            self.text.set_content(Global_logger.return_logs())
            def update_logs(logs):
                self.text.set_content(logs)
                scroll_area.scroll_to(percent=1.0)
            Global_logger.register_callback(update_logs)
            self.text.props('readonly')
        ui.separator().style("flex-grow:1;")
        with ui.row().style('margin-top:20px; justify-content: flex-start; gap: 20px;'):
            ui.button('清空日志', on_click=Global_logger.clear).style('font-size: 16px; padding: 10px 20px;')
            ui.button('导出日志', on_click=Global_logger.get_all).style('font-size: 16px; padding: 10px 20px;')
        ui.separator().style("flex-grow:1;")
        ui.button('关闭应用', on_click=self._close_app).style('position: fixed; bottom: 20px; right: 20px; font-size: 16px; padding: 10px 20px;')
        ui.run(reload=False, title='ISBN查重系统')

    def _switch_database(self, db_path: str):
        self.database = ISBN_Database(db_path)
        Global_logger.append(f"已切换数据库：{db_path}")
        ui.navigate.to('/')

    def _on_single_check(self, value: str) -> None:
        SearchEngine.search_single(self.database, value)

    def _on_batch_check(self, value: str) -> None:
        isbn_list = value.splitlines()
        SearchEngine.search_batch(self.database, isbn_list)

    async def _send_to_query(self, event) -> None:
        files = getattr(event, 'files', None) or [getattr(event, 'file', None)]
        files = [f for f in files if f]
        await SearchEngine.batch_clean(self.database, files)

    async def _send_to_update(self, event) -> None:
        await self.database.update_database(single=event.file)

    async def _setup_database(self, event: UploadEventArguments):
        await self.database.create_database(single=event.file)

    async def _setup_database_multi(self, event: MultiUploadEventArguments):
        await self.database.create_database(multi=event.files)

    async def _close_app(self):
        if Global_logger:
            Global_logger.clear()
        import asyncio
        ui.run_javascript('window.close();')
        await asyncio.sleep(1.5)
        try:
            import signal
            os.kill(os.getpid(), signal.SIGINT)
        except Exception:
            os._exit(0)

if getattr(sys, 'frozen', False):
    meipass = getattr(sys, '_MEIPASS', None)
    if meipass:
        candidate = os.path.join(meipass, 'main.py')
        if os.path.exists(candidate):
            sys.argv[0] = candidate
    else:
        try:
            import inspect
            src = inspect.getsource(sys.modules['__main__'])
            fd, path = tempfile.mkstemp(suffix='.py', text=True)
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                f.write(src)
            sys.argv[0] = path
        except Exception:
            pass

if __name__ == "__main__":
    app = Aplication()
    if not app.database.is_exist():
        app._build_init_ui()
        Global_logger.append("初始化中...请上传文件创建数据库")
    else:
        app._build_main_ui()
        Global_logger.append("数据库已存在，进入主界面")
