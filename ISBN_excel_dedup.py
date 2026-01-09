# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox
import sqlite3
import os
from openpyxl import load_workbook

def fuzzy_isbn_col(header_row):
    for idx, cell in enumerate(header_row):
        val = str(cell.value).lower() if cell.value else ""
        if any(x in val for x in ["isbn", "书号"]):
            return idx
    return None

def get_db_isbns(db_path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT isbn FROM isbn_records")
    db_isbns = set(row[0] for row in cur.fetchall())
    conn.close()
    return db_isbns

def dedup_excel(file_path, db_path, progress_callback=None):
    wb = load_workbook(file_path)
    ws = wb.active
    header = [cell for cell in ws[1]]
    isbn_col = fuzzy_isbn_col(header)
    if isbn_col is None:
        raise Exception("未找到ISBN列，请检查表头")
    db_isbns = get_db_isbns(db_path)
    orig_rows = ws.max_row - 1
    repeat_rows = 0
    keep_rows = orig_rows
    # 逆序删除
    for row_idx in range(ws.max_row, 1, -1):
        cell = ws.cell(row=row_idx, column=isbn_col+1)
        isbn = str(cell.value).strip() if cell.value else ""
        if isbn in db_isbns:
            ws.delete_rows(row_idx)
            repeat_rows += 1
            keep_rows -= 1
        if progress_callback:
            progress_callback(orig_rows, repeat_rows, keep_rows)
    # 输出副本
    out_path = file_path.replace('.xlsx', '_dedup.xlsx')
    wb.save(out_path)
    return orig_rows, repeat_rows, keep_rows, out_path

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel/CSV ISBN去重工具")
        self.file_path = ""
        self.db_path = "isbn_records.db"
        self.progress_var = tk.StringVar()
        self.result_var = tk.StringVar()
        self.setup_ui()

    def setup_ui(self):
        frm = tk.Frame(self.root)
        frm.pack(padx=20, pady=20)
        tk.Label(frm, text="选择Excel文件:").grid(row=0, column=0, sticky='e')
        tk.Button(frm, text="浏览", command=self.select_file).grid(row=0, column=1)
        self.file_label = tk.Label(frm, text="未选择文件", width=40, anchor='w')
        self.file_label.grid(row=0, column=2)
        tk.Button(frm, text="开始去重", command=self.run_dedup).grid(row=1, column=1, pady=10)
        tk.Label(frm, textvariable=self.progress_var, fg='blue').grid(row=2, column=0, columnspan=3)
        tk.Label(frm, textvariable=self.result_var, fg='green').grid(row=3, column=0, columnspan=3)

    def select_file(self):
        filetypes = [("Excel文件", "*.xlsx")]
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))

    def run_dedup(self):
        if not self.file_path:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        if not os.path.exists(self.db_path):
            messagebox.showerror("错误", f"数据库文件不存在: {self.db_path}")
            return
        self.progress_var.set("处理中...")
        self.root.update()
        try:
            orig, repeat, keep, out_path = dedup_excel(
                self.file_path, self.db_path,
                progress_callback=lambda o, r, k: self.progress_var.set(f"原始行数: {o}，已删除: {r}，剩余: {k}")
            )
            self.result_var.set(f"处理完成！\n原始行数: {orig}\n重复行数: {repeat}\n输出行数: {keep}\n输出文件: {os.path.basename(out_path)}")
        except Exception as e:
            messagebox.showerror("处理失败", str(e))
            self.progress_var.set("")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
